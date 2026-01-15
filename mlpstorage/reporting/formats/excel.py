"""
Excel format handler for MLPerf Storage benchmark results.

Provides Excel workbook export with analysis and pivot tables.
"""

import io
from typing import List, Dict, Any, Optional, TYPE_CHECKING

from mlpstorage.reporting.formats import ReportFormat, FormatRegistry, FormatConfig
from mlpstorage.config import PARAM_VALIDATION

if TYPE_CHECKING:
    from mlpstorage.reporting import Result

# Excel dependencies are optional
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.worksheet.table import Table, TableStyleInfo
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


@FormatRegistry.register
class ExcelFormat(ReportFormat):
    """Format results as Excel workbooks with analysis and pivot tables."""

    # Cell styles
    HEADER_FILL = PatternFill(start_color='366092', end_color='366092', fill_type='solid') if EXCEL_AVAILABLE else None
    HEADER_FONT = Font(color='FFFFFF', bold=True) if EXCEL_AVAILABLE else None
    CLOSED_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid') if EXCEL_AVAILABLE else None
    OPEN_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid') if EXCEL_AVAILABLE else None
    INVALID_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid') if EXCEL_AVAILABLE else None

    def __init__(self, config: Optional[FormatConfig] = None):
        """
        Initialize the Excel formatter.

        Args:
            config: Optional configuration.
        """
        super().__init__(config)
        if not EXCEL_AVAILABLE:
            pass  # Will raise error on generate

    @property
    def name(self) -> str:
        return "excel"

    @property
    def extension(self) -> str:
        return "xlsx"

    @property
    def content_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def _check_available(self):
        """Check if Excel support is available."""
        if not EXCEL_AVAILABLE:
            raise ImportError(
                "Excel support requires openpyxl. "
                "Install with: pip install openpyxl"
            )

    def _get_category_fill(self, category: PARAM_VALIDATION):
        """Get fill color for a category."""
        if category == PARAM_VALIDATION.CLOSED:
            return self.CLOSED_FILL
        elif category == PARAM_VALIDATION.OPEN:
            return self.OPEN_FILL
        return self.INVALID_FILL

    def _style_header_row(self, ws, num_cols: int):
        """Apply styling to header row."""
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center')

    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _add_summary_sheet(self, wb: 'Workbook', results: List['Result'],
                           workload_results: Dict) -> None:
        """Add summary sheet with counts and categories."""
        ws = wb.active
        ws.title = "Summary"

        # Counts
        closed = sum(1 for r in results if r.category == PARAM_VALIDATION.CLOSED)
        open_count = sum(1 for r in results if r.category == PARAM_VALIDATION.OPEN)
        invalid = sum(1 for r in results if r.category == PARAM_VALIDATION.INVALID)

        # Title
        ws['A1'] = "MLPerf Storage Benchmark Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')

        # Summary stats
        ws['A3'] = "Summary Statistics"
        ws['A3'].font = Font(bold=True)

        ws['A4'] = "Total Runs:"
        ws['B4'] = len(results)

        ws['A5'] = "CLOSED Runs:"
        ws['B5'] = closed
        ws['B5'].fill = self.CLOSED_FILL

        ws['A6'] = "OPEN Runs:"
        ws['B6'] = open_count
        ws['B6'].fill = self.OPEN_FILL

        ws['A7'] = "INVALID Runs:"
        ws['B7'] = invalid
        ws['B7'].fill = self.INVALID_FILL

        ws['A9'] = "Workload Submissions:"
        ws['B9'] = len(workload_results)

        # Category breakdown by type
        ws['A11'] = "Breakdown by Benchmark Type"
        ws['A11'].font = Font(bold=True)

        type_counts = {}
        for r in results:
            btype = r.benchmark_type.value if r.benchmark_type else 'unknown'
            if btype not in type_counts:
                type_counts[btype] = {'closed': 0, 'open': 0, 'invalid': 0}

            if r.category == PARAM_VALIDATION.CLOSED:
                type_counts[btype]['closed'] += 1
            elif r.category == PARAM_VALIDATION.OPEN:
                type_counts[btype]['open'] += 1
            else:
                type_counts[btype]['invalid'] += 1

        row = 12
        headers = ['Benchmark Type', 'CLOSED', 'OPEN', 'INVALID', 'Total']
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
            ws.cell(row=row, column=col).font = Font(bold=True)

        row += 1
        for btype, counts in type_counts.items():
            ws.cell(row=row, column=1, value=btype)
            ws.cell(row=row, column=2, value=counts['closed'])
            ws.cell(row=row, column=2).fill = self.CLOSED_FILL
            ws.cell(row=row, column=3, value=counts['open'])
            ws.cell(row=row, column=3).fill = self.OPEN_FILL
            ws.cell(row=row, column=4, value=counts['invalid'])
            ws.cell(row=row, column=4).fill = self.INVALID_FILL
            ws.cell(row=row, column=5, value=sum(counts.values()))
            row += 1

        self._auto_adjust_columns(ws)

    def _add_runs_sheet(self, wb: 'Workbook', results: List['Result']) -> None:
        """Add detailed runs sheet."""
        ws = wb.create_sheet("Runs")

        headers = ['Run ID', 'Benchmark Type', 'Model', 'Command', 'Category',
                  'Processes', 'Accelerator', 'AU %', 'Throughput', 'Issues']

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        self._style_header_row(ws, len(headers))

        for row_idx, result in enumerate(results, 2):
            run = result.benchmark_run
            metrics = result.metrics or {}

            au_pct = metrics.get('au', metrics.get('accelerator_utilization', ''))
            throughput = metrics.get('throughput', '')

            ws.cell(row=row_idx, column=1, value=str(run.run_id) if hasattr(run, 'run_id') else '')
            ws.cell(row=row_idx, column=2, value=result.benchmark_type.value if result.benchmark_type else '')
            ws.cell(row=row_idx, column=3, value=str(result.benchmark_model) if result.benchmark_model else '')
            ws.cell(row=row_idx, column=4, value=result.benchmark_command or '')

            category_cell = ws.cell(row=row_idx, column=5, value=result.category.value.upper() if result.category else '')
            category_cell.fill = self._get_category_fill(result.category)

            ws.cell(row=row_idx, column=6, value=run.num_processes if hasattr(run, 'num_processes') else '')
            ws.cell(row=row_idx, column=7, value=str(run.accelerator) if hasattr(run, 'accelerator') and run.accelerator else '')
            ws.cell(row=row_idx, column=8, value=au_pct if isinstance(au_pct, (int, float)) else '')
            ws.cell(row=row_idx, column=9, value=throughput if isinstance(throughput, (int, float)) else '')
            ws.cell(row=row_idx, column=10, value=len(result.issues))

        # Create table
        if len(results) > 0:
            table_ref = f"A1:{get_column_letter(len(headers))}{len(results) + 1}"
            table = Table(displayName="RunsTable", ref=table_ref)
            style = TableStyleInfo(
                name="TableStyleMedium9", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=False
            )
            table.tableStyleInfo = style
            ws.add_table(table)

        self._auto_adjust_columns(ws)

    def _add_workloads_sheet(self, wb: 'Workbook', workload_results: Dict) -> None:
        """Add workload submissions sheet."""
        if not workload_results:
            return

        ws = wb.create_sheet("Workloads")

        headers = ['Model', 'Accelerator', 'Benchmark Type', 'Category', 'Runs', 'Issues']

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        self._style_header_row(ws, len(headers))

        for row_idx, ((model, accelerator), result) in enumerate(workload_results.items(), 2):
            num_runs = len(result.benchmark_run) if isinstance(result.benchmark_run, list) else 1

            ws.cell(row=row_idx, column=1, value=str(model) if model else '')
            ws.cell(row=row_idx, column=2, value=str(accelerator) if accelerator else '')
            ws.cell(row=row_idx, column=3, value=result.benchmark_type.value if result.benchmark_type else '')

            category_cell = ws.cell(row=row_idx, column=4, value=result.category.value.upper() if result.category else '')
            category_cell.fill = self._get_category_fill(result.category)

            ws.cell(row=row_idx, column=5, value=num_runs)
            ws.cell(row=row_idx, column=6, value=len(result.issues))

        self._auto_adjust_columns(ws)

    def _add_metrics_sheet(self, wb: 'Workbook', results: List['Result']) -> None:
        """Add performance metrics sheet."""
        ws = wb.create_sheet("Metrics")

        # Collect all metric keys
        all_keys = set()
        for result in results:
            if result.metrics:
                all_keys.update(result.metrics.keys())

        # Prioritize common metrics
        priority_keys = ['au', 'throughput', 'train_au_percentage',
                        'train_throughput_samples_per_second', 'samples_processed']
        ordered_keys = [k for k in priority_keys if k in all_keys]
        ordered_keys.extend(k for k in sorted(all_keys) if k not in priority_keys)

        if not ordered_keys:
            ws['A1'] = "No metrics available"
            return

        headers = ['Run ID', 'Model', 'Category'] + ordered_keys

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        self._style_header_row(ws, len(headers))

        for row_idx, result in enumerate(results, 2):
            run = result.benchmark_run
            metrics = result.metrics or {}

            ws.cell(row=row_idx, column=1, value=str(run.run_id) if hasattr(run, 'run_id') else '')
            ws.cell(row=row_idx, column=2, value=str(result.benchmark_model) if result.benchmark_model else '')

            category_cell = ws.cell(row=row_idx, column=3, value=result.category.value.upper() if result.category else '')
            category_cell.fill = self._get_category_fill(result.category)

            for col_idx, key in enumerate(ordered_keys, 4):
                value = metrics.get(key, '')
                if isinstance(value, (list, dict)):
                    value = str(value)
                ws.cell(row=row_idx, column=col_idx, value=value)

        self._auto_adjust_columns(ws)

    def _add_chart(self, wb: 'Workbook', results: List['Result']) -> None:
        """Add performance chart to metrics sheet."""
        if 'Metrics' not in wb.sheetnames:
            return

        ws = wb['Metrics']

        # Find AU column
        au_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value in ('au', 'train_au_percentage'):
                au_col = col
                break

        if au_col is None or ws.max_row < 2:
            return

        # Create bar chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Accelerator Utilization by Run"
        chart.y_axis.title = "AU %"
        chart.x_axis.title = "Run"

        data = Reference(ws, min_col=au_col, min_row=1, max_row=ws.max_row)
        categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.shape = 4
        chart.width = 15
        chart.height = 10

        ws.add_chart(chart, f"{get_column_letter(ws.max_column + 2)}2")

    def _add_cluster_info_sheet(self, wb: 'Workbook', cluster_info: Dict) -> None:
        """Add cluster information sheet."""
        if not cluster_info:
            return

        ws = wb.create_sheet("Cluster Info")

        # Summary info
        ws['A1'] = "Cluster Summary"
        ws['A1'].font = Font(size=14, bold=True)

        ws['A3'] = "Total Hosts:"
        ws['B3'] = cluster_info.get('num_hosts', 0)

        ws['A4'] = "Total Memory (GB):"
        ws['B4'] = round(cluster_info.get('total_memory_bytes', 0) / (1024**3), 1)

        ws['A5'] = "Total Cores:"
        ws['B5'] = cluster_info.get('total_cores', 0)

        # Host details
        hosts = cluster_info.get('hosts', [])
        if hosts:
            ws['A7'] = "Host Details"
            ws['A7'].font = Font(size=14, bold=True)

            headers = ['Hostname', 'CPU Model', 'Cores', 'Memory (GB)']
            for col, header in enumerate(headers, 1):
                ws.cell(row=8, column=col, value=header)
                ws.cell(row=8, column=col).font = Font(bold=True)

            for row_idx, host in enumerate(hosts, 9):
                ws.cell(row=row_idx, column=1, value=host.get('hostname', ''))

                cpu = host.get('cpu', {})
                ws.cell(row=row_idx, column=2, value=cpu.get('model', '') if cpu else '')
                ws.cell(row=row_idx, column=3, value=cpu.get('num_cores', '') if cpu else '')

                memory = host.get('memory', {})
                mem_gb = memory.get('total', 0) / (1024**3) if memory else 0
                ws.cell(row=row_idx, column=4, value=round(mem_gb, 1))

        self._auto_adjust_columns(ws)

    def _add_param_ranges_sheet(self, wb: 'Workbook', param_ranges: Dict[str, Dict]) -> None:
        """Add parameter ranges sheet."""
        if not param_ranges:
            return

        ws = wb.create_sheet("Parameter Ranges")

        headers = ['Workload', 'Parameter', 'Min', 'Max', 'Avg', 'Values Count']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        self._style_header_row(ws, len(headers))

        row = 2
        for workload, ranges in param_ranges.items():
            for param, stats in ranges.items():
                if isinstance(stats, dict):
                    ws.cell(row=row, column=1, value=workload)
                    ws.cell(row=row, column=2, value=param)
                    ws.cell(row=row, column=3, value=stats.get('min', ''))
                    ws.cell(row=row, column=4, value=stats.get('max', ''))
                    ws.cell(row=row, column=5, value=stats.get('avg', ''))
                    ws.cell(row=row, column=6, value=len(stats.get('values', [])))
                    row += 1

        self._auto_adjust_columns(ws)

    def generate(self, results: List['Result'], workload_results: Dict,
                 advanced_data: Optional[Dict] = None) -> bytes:
        """
        Generate Excel workbook.

        Args:
            results: List of Result objects.
            workload_results: Dictionary of workload results.
            advanced_data: Optional advanced data.

        Returns:
            Excel file content as bytes.
        """
        self._check_available()

        wb = Workbook()

        # Add sheets
        self._add_summary_sheet(wb, results, workload_results)
        self._add_runs_sheet(wb, results)
        self._add_workloads_sheet(wb, workload_results)
        self._add_metrics_sheet(wb, results)
        self._add_chart(wb, results)

        # Add advanced sheets if data provided
        if advanced_data:
            if 'cluster_info' in advanced_data:
                self._add_cluster_info_sheet(wb, advanced_data['cluster_info'])
            if 'param_ranges' in advanced_data:
                self._add_param_ranges_sheet(wb, advanced_data['param_ranges'])

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.read()

    def create_workbook(self, results: List['Result'], workload_results: Dict,
                        advanced_data: Optional[Dict] = None) -> 'Workbook':
        """
        Create Excel workbook object.

        Args:
            results: List of Result objects.
            workload_results: Dictionary of workload results.
            advanced_data: Optional advanced data.

        Returns:
            Workbook object.
        """
        self._check_available()

        wb = Workbook()

        self._add_summary_sheet(wb, results, workload_results)
        self._add_runs_sheet(wb, results)
        self._add_workloads_sheet(wb, workload_results)
        self._add_metrics_sheet(wb, results)
        self._add_chart(wb, results)

        if advanced_data:
            if 'cluster_info' in advanced_data:
                self._add_cluster_info_sheet(wb, advanced_data['cluster_info'])
            if 'param_ranges' in advanced_data:
                self._add_param_ranges_sheet(wb, advanced_data['param_ranges'])

        return wb
